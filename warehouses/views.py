from django.shortcuts import render, redirect,get_object_or_404
from django.core.paginator import Paginator
from django.contrib import messages
from django.urls import reverse
from django.db.models import ProtectedError,Q,Sum,Count
from django.utils.datastructures import MultiValueDictKeyError
from .models import Power, Mark, Engine
from .filters import (InventoryFilter,UnusableFilter,NewPumpFilter,EngineFilter,GeneralEngineFilter,
    DebtSituationFilter,PumpFilter,OrderFilter,SeconhandFilter,WorkshopExitSlipFilter)
from .forms import *
from account.decorators import administrator,admin
from other_materials.forms import CategoryStockOut2Form
from hydrophore.models import OutboundWorkOrder,WorkshopExit,RepairReturn
from django.http import JsonResponse,HttpResponse
from django.template.loader import render_to_string
from openpyxl import Workbook

def handle_deletion(request, model, object_id, redirect_url, success_message, error_message, protected_error_message):
    try:
        obj = model.objects.get(id=object_id)
        obj.delete() 
        messages.success(request, success_message.format(obj))
    except model.DoesNotExist:
        messages.warning(request, error_message)
    except ProtectedError as e:
        model_name = e.args[0].split(' ')[-1] 
        messages.warning(request, protected_error_message.format(obj, model_name))
    return redirect(redirect_url)

def paginate_items(request, items, per_page=10):
    paginator = Paginator(items, per_page)  
    page_number = request.GET.get('page') 
    page_obj = paginator.get_page(page_number)  

    return page_obj

def engine_locations_update(id,location):
    item =Engine.objects.get(id=id)
    item.location = location
    item.save()

@admin
def newMark(request):
    if request.method == "POST":
        form = MarkForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f"*{form.data['engine_mark']}* Kayıt başarılı.")
            return redirect('add_mark')
        else:
            messages.warning(request, form.errors.as_ul())
    else:
        form = MarkForm()
    context = {
        'items': Mark.objects.all().order_by("-id"),
        'form': form
    }
    return render(request, "new_mark.html", context)

@administrator
def markDelete(request, myid):
    return handle_deletion(
        request,
        Mark,
        myid,
        'add_mark',
        "*{0}* Markası başarıyla silindi.",
        "Marka bulunamadı.",
        "*{0}* Marka kaydı {1} tabloda kullanılıyor, silinemez."
    )

@admin
def newPower(request):
    if request.method == "POST":
        form = PowerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f"*{form.data['engine_power']}* Kayıt başarılı.")
            return redirect('add_power')
        else:
            messages.warning(request, form.errors.as_ul())
    else:
        form = PowerForm()

    context = {
        'items': Power.objects.all().order_by("-id"),
        'form': form
    }
    return render(request, "new_power.html", context)

@administrator
def powerDelete(request, myid):
    return handle_deletion(
        request,
        Power,
        myid,
        'add_power',
        "*{0}* Güç değeri başarıyla silindi.",
        "Güç değeri bulunamadı.",
        "*{0}* Güç kaydı {1} tabloda kullanılıyor, silinemez."
    )

@administrator
def engineDelete(request, myid):
    redirect_url = request.GET.get('next', 'engine_homepage')
    try:
        obj = Engine.objects.get(id=myid)

        if Inventory.objects.filter(engine=obj).exists():
            messages.warning(request, f"*{obj}* Motor kaydı Kuyu depoda kullanılıyor, silinemez.")
            return redirect(redirect_url)

        return handle_deletion(
            request,
            Engine,
            myid,
            redirect_url,
            "*{0}* Motor kaydı başarıyla silindi.",
            "Motor kaydı bulunamadı.",
            "*{0}* Motor kaydı {1} tabloda kullanılıyor, silinemez."
        )

    except Engine.DoesNotExist:
        messages.warning(request, "Motor kaydı bulunamadı.")
        return redirect(redirect_url)

def engine_homepage(request):
    # Tüm motorlar
    engine_list = Engine.objects.all().order_by("-id")
    
    # Filtreleme (opsiyonel)
    engine_filter = EngineFilter(request.GET, queryset=engine_list)
    filtered_qs = engine_filter.qs

    # Pagination
    page_obj = paginate_items(request, filtered_qs)

    # Lokasyon bazlı sayılar
    location_counts = (
        Engine.objects.values('location')
        .annotate(count=Count('id'))
        .order_by('location')
    )
    # Lokasyonları sözlük haline getirme
    location_dict = {str(item['location']): item['count'] for item in location_counts}

    context = {
        'well': location_dict.get("1", 0),
        'repair': Repair.objects.filter(engine__location= "2").count(),
        'secondhand': location_dict.get("3", 0),
        'unusable': location_dict.get("4", 0),
        'new': location_dict.get("5", 0),
        'contractor' : location_dict.get("6", 0),
        'total': filtered_qs.count(),
        'items': page_obj,
        'query_string': request.GET.urlencode(),
        'filter': engine_filter,  # template’de filtre formunu göstermek için
    }

    return render(request, "engine_homepage.html", context)

@admin
def engine_edit(request, pk):
    engine = get_object_or_404(Engine, pk=pk)

    # next'i al
    next_url = request.GET.get('next')

    if request.method == 'POST':
        form = EngineForm(request.POST, instance=engine)

        # POST'tan tekrar al
        next_url = request.POST.get('next')

        if form.is_valid():
            form.save()
            messages.success(request, "Motor bilgileri güncellendi.")
            return redirect(next_url or 'engine_homepage')
        else:
            messages.warning(request, form.errors.as_ul())
    else:
        form = EngineForm(instance=engine)

    return render(request, 'new_engine.html', {
        'form': form,
        'next': next_url
    })

@admin
def newPump(request):
    if request.method == "POST":
        form = PumpForm(request.POST)
        if form.is_valid():
            item = form.save()
            messages.success(
                request,
                f'*{ item }* Pompa kaydı başarılı bir şekilde oluşturuldu.'
            )
            return redirect('add_pump')  
            
        else:
            messages.warning(
                request,
                f"Formda hatalar var. Lütfen kontrol edin: {form.errors.as_ul()}"
            )
    else:
        form = PumpForm()  
    context = {
        'form': form
    }
    return render(request, "new_pump.html", context)

@administrator
def pumpDelete(request, myid):
    return handle_deletion(
        request,
        Pump,
        myid,
        'pump_homepage',
        "*{0}* Pompa kaydı başarıyla silindi.",
        "Pompa kaydı bulunamadı.",
        "*{0}* Pompa kaydı {1} tabloda kullanılıyor, silinemez."
    )

@administrator
def pump_edit(request, pk):
    pump = get_object_or_404(Pump, pk=pk)
    if request.method == 'POST':
        form = PumpEditForm(request.POST, instance=pump)
        if form.is_valid():
            form.save()
            messages.success(request, f"Pompa bilgileri güncellendi.")
            return redirect('pump_homepage')
    else:
        form = PumpEditForm(instance=pump)
    return render(request, 'new_pump.html', {'form': form})

def pump_homepage(request):
    pump_list = Pump.objects.all().order_by("-id")

    # 🔍 Filtreleme
    pump_filter = PumpFilter(request.GET, queryset=pump_list)
    filtered_qs = pump_filter.qs

    # Pagination
    page_obj = paginate_items(request, filtered_qs)

    context = {
        'well': Inventory.objects.filter(pump__isnull=False).count(),
        'repair': Repair.objects.filter(pump__isnull=False).count(),
        'secondhand': Seconhand.objects.filter(pump__isnull=False).count(),
        'unusable': Unusable.objects.filter(pump__isnull=False).count(),
        'new': NewWarehousePump.objects.aggregate(total=Sum('quantity'))['total'] or 0,
        'total': filtered_qs.count(),
        'items': page_obj,
        'query_string': request.GET.urlencode(),
        'filter': pump_filter,
    }
    return render(request, "pump_homepage.html", context)

def inventory_homepage(request):
    queryset = Inventory.objects.all().order_by("id")
    inventory_filter = InventoryFilter(request.GET, queryset=queryset)
    filtered_qs = inventory_filter.qs
    page_obj = paginate_items(request, filtered_qs)

    if request.method == 'POST':
        form = NewInventoryForm(request.POST)
        if form.is_valid(): 
            obj = form.save(commit=False)
            obj.status = "passive"
            obj.save()
            messages.success(request, f"*{obj.well_number}* Yeni kuyu kaydı başarıyla eklendi. ")
            return redirect('inventory')
        else:
            messages.warning(request, form.errors.as_ul())
    else:
        form = NewInventoryForm()
        
    context = {
        'filter': inventory_filter,
        'items': page_obj,
        'total': filtered_qs.count(),
        'query_string': request.GET.urlencode(),
        'form' : form
    }
    return render(request, "inventory_homepage.html", context)

@admin
def inventory_edit(request, pk):
    inventory = get_object_or_404(Inventory, pk=pk)
    next_url = request.GET.get('next')
    if inventory.status == "passive":
        messages.info(request, f"*{inventory.well_number}* Kuyu PASİF durumda. Aktif etmek için kuyuya motor ve pompa ekleyin.")
        return redirect(next_url or 'inventory')

    if request.method == 'POST':
        # POST'tan tekrar al
        next_url = request.POST.get('next')
        form = InventoryForm(request.POST, instance=inventory)
        engine_form = EngineForm(request.POST,instance=inventory.engine)
        pump_form = PumpForm(request.POST)
        if form.is_valid() and engine_form.is_valid():
            item = form.save(commit=False)
            engine_form.save()
            existing_pump = Pump.objects.filter(
                    pump_type=pump_form.data["pump_type"],
                    pump_breed=pump_form.data["pump_breed"],
                    pump_mark=pump_form.data["pump_mark"]
                ).first()

            if existing_pump:
                pump = existing_pump
            else:
                pump = pump_form.save()
            item.pump = pump
            item.save()
            messages.success(request, "Kuyu bilgileri güncellendi.")
            return redirect(next_url or 'inventory')
        else:
            messages.warning(request, form.errors.as_ul())
            messages.warning(request, engine_form.errors.as_ul())
    else:
        engine_form = EngineForm(instance=inventory.engine)
        pump_form = PumpForm(instance=inventory.pump)
        form = InventoryForm(instance=inventory)
    return render(request, 'new_inventory.html', {
        'form': form,
        'engine_form':engine_form,
        'pump_form':pump_form,
        'next': next_url
    })

@admin
def add_inventory(request):
    next_url = request.GET.get('next')
    if request.method == 'POST':
        next_url = request.GET.get('next')
        form = InventoryForm(request.POST or None)
        engine_form = EngineForm(request.POST or None)
        pump_form = PumpForm(request.POST or None)
        if form.is_valid() and engine_form.is_valid(): 
            existing_pump = Pump.objects.filter(
                    pump_type=pump_form.data["pump_type"],
                    pump_breed=pump_form.data["pump_breed"],
                    pump_mark=pump_form.data["pump_mark"]
                ).first()

            if existing_pump:
                pump = existing_pump
            else:
                pump = pump_form.save()
            engine = engine_form.save(commit=False)
            engine.location = "1"
            engine.save()
            item = form.save(commit=False)
            item.engine = engine
            item.pump = pump
            item.save()
            messages.success(request, f"*{item.well_number}* Kuyu kaydı başarıyla eklendi ve ilgili alanlar güncellendi.")
            return redirect('inventory')
        else:
            messages.warning(request, form.errors.as_ul())
            messages.warning(request, engine_form.errors.as_ul())
    else:
        engine_form = EngineForm()
        pump_form = PumpForm()
        form = InventoryForm()

    context = {'form': form ,'engine_form':engine_form,'pump_form':pump_form,'next': next_url}
    return render(request, 'new_inventory.html', context)

@admin
def export_inventory(request):
    # queryset ve filtreleme
    inventory_list = Inventory.objects.select_related('engine', 'pump').all()

    inventory_filter = InventoryFilter(request.GET, queryset=inventory_list)
    filtered_inventory_list = inventory_filter.qs
   
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Kuyular")

    # Excel başlıkları – Kuyu odaklı
    ws.append([
        "Kuyu Numarası",
        "İlçe",
        "Adres",
        "Demontaj Derinliği",
        "Montaj Derinliği",
        "Depo Bilgisi",
        "Boru Tipi",
        "Kablo",
        "Motor",
        "Pompa",
        "Debi",
        "Açıklama",
        "Durum",
        "Oluşturulma Tarihi",
        "Güncellenme Tarihi"
    ])

    # verileri iterator ile yaz
    for obj in filtered_inventory_list.iterator(chunk_size=1000):
        ws.append([
            obj.well_number,
            obj.get_district_display(),
            obj.address,
            obj.disassembly_depth,
            obj.mounting_depth,
            obj.tank_info,
            obj.pipe_type,
            obj.cable,
            str(obj.engine) if obj.engine else "",
            str(obj.pump) if obj.pump else "",
            obj.flow,
            obj.comment,
            obj.get_status_display(),
            obj.created_at.strftime("%Y-%m-%d %H:%M") if obj.created_at else "",
            obj.updated_at.strftime("%Y-%m-%d %H:%M") if obj.updated_at else ""
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="kuyu_bilgileri.xlsx"'

    wb.save(response)
    return response

@administrator
def delete_inventory(request, id):
    return handle_deletion(
        request,
        Inventory,
        id,
        'inventory',
        "*{0}* Nolu kuyu kaydı başarıyla silindi.",
        "Kuyu kaydı bulunamadı.",
        "*{0}* Kuyu kaydı {1} tabloda kullanılıyor, silinemez."
    )

def seconhand(request):
    queryset = Seconhand.objects.filter(
        Q(engine__isnull=False) | Q(pump__isnull=False)
    ).select_related('engine','pump').order_by('row_identifier')

    # 🔍 FILTER
    seconhand_filter = SeconhandFilter(request.GET, queryset=queryset)

    # 📄 PAGINATION
    page_obj = paginate_items(request, seconhand_filter.qs)

    engine_form = EngineForm(request.POST or None)
    pump_form = PumpForm(request.POST or None)

    if request.method == 'POST':
        secondhand_id = request.POST.get("secondhand_row")

        if secondhand_id:
            item = Seconhand.objects.get(id=secondhand_id)
            engine_valid = False
            if engine_form.is_valid():
                engine = engine_form.save(commit=False)
                engine.location = "3"
                engine.save()
                item.engine = engine
                engine_valid = True
            try:
                pump_valid = False
                existing_pump = Pump.objects.filter(
                    pump_type=pump_form.data["pump_type"],
                    pump_breed=pump_form.data["pump_breed"],
                    pump_mark=pump_form.data["pump_mark"]
                ).first()

                if existing_pump:
                    item.pump = existing_pump
                else:
                    item.pump = pump_form.save()
                pump_valid = True
            except MultiValueDictKeyError:
                pass
            
            if engine_valid or pump_valid:
                item.save()
                messages.success(request,f"{item.row_identifier} Sırası güncellendi.")
                return redirect('seconhand')
            
        messages.warning(
            request,
            engine_form.errors.as_ul() or pump_form.errors.as_ul()
        )

    context = {
        'engine_form': engine_form,
        'pump_form': pump_form,

        'items': page_obj,
        'filter': seconhand_filter,

        'null_list': Seconhand.objects.filter(
            Q(engine__isnull=True) | Q(pump__isnull=True)
        ).select_related('engine','pump').order_by("row_identifier"),
        'total': seconhand_filter.qs.count(),
        'query_string': request.GET.urlencode(),
    }

    return render(request, "secondhand_page.html", context)

@administrator
def delete_seconhand_pump(request, id):
    seconhand = get_object_or_404(Seconhand,id=id)
    messages.success(request,f"{seconhand.row_identifier} sırasındaki {str(seconhand.pump)} pompa kaydı silindi.")
    seconhand.pump = None
    seconhand.save()
    return redirect("seconhand")

@admin
def export_seconhand(request):
    # Filtreleri uygula
    order_list = Seconhand.objects.select_related('pump', 'engine').all()
    order_filter = SeconhandFilter(request.GET, queryset=order_list)
    filtered_order_list = order_filter.qs

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="2.El Depo")

    # Başlıklar
    ws.append([
        "Sıra",
        "Pompa",
        "Motor",
    ])

    # Verileri yaz
    for obj in filtered_order_list.iterator(chunk_size=1000):
        ws.append([
            obj.row_identifier,
            str(obj.pump )if obj.pump else "",
            str(obj.engine) if obj.engine else "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="2_el_depo.xlsx"'

    wb.save(response)
    return response

def warehouses_repair(request):
    repair_list = Repair.objects.filter(status="active")
    page_obj = paginate_items(request, repair_list)

    contex = {
        'total': repair_list.count(),  
        'items': page_obj, 
        'query_string': request.GET.urlencode(),
    }
    return render(request, "repair_page.html",contex)

def repair(request):
    repair_list = Repair.objects.filter(status="active")
    page_obj = paginate_items(request, repair_list)

    contex = {
        'total': repair_list.count(),  
        'items': page_obj, 
        'query_string': request.GET.urlencode(),
    }
    return render(request, "repair_page.html",contex)

def all_repair(request):
    repair_list = Repair.objects.filter(status="passive")
    page_obj = paginate_items(request, repair_list)

    contex = {
        'total': repair_list.count(),  
        'items': page_obj, 
        'query_string': request.GET.urlencode(),
    }
    return render(request, "repair_page.html",contex)

@admin
def repair_edit(request, id):
    repair = get_object_or_404(Repair, order__pk=id)
    if repair.engine:
        engine = Engine.objects.filter(location="6",engine_type=repair.engine.engine_type,engine_power=repair.engine.engine_power)
    else:
        engine =None
    row_identifier = Seconhand.objects.filter(engine__isnull=True,pump__isnull=True)
    next_url = request.GET.get('next')
    if request.method == "POST":
        next_url = request.POST.get('next')
        engine_choice = request.POST.get("engine")
        pump_choice = request.POST.get("pump")

        engine_row_id = request.POST.get("engine_secondhand_row")
        pump_row_id = request.POST.get("pump_secondhand_row")
        engine_contractor_row_id = request.POST.get("engine_contractor")

        # -------- PERT --------
        if engine_choice == "unusable" or pump_choice == "unusable":
            Unusable.objects.create(
                well_number=repair.order.inventory,
                engine=repair.engine if engine_choice == "unusable" else None,
                pump=repair.pump if pump_choice == "unusable" else None
            )
            repair.engine_info = 'Pert Depo' if engine_choice == "unusable" else None
            repair.pump_info = 'Pert Depo' if pump_choice == "unusable" else None
            
            if engine_choice == "unusable" and repair.engine:
                engine_locations_update(repair.engine.id, "4")

        # -------- 2.EL --------
        if engine_choice == "secondhand" and engine_row_id:
            row = Seconhand.objects.get(id=engine_row_id)
            row.engine = repair.engine
            row.save()
            repair.engine_info = f"2.El Depo | {row.row_identifier}"
            if repair.engine:
                engine_locations_update(repair.engine.id, "3")

        if pump_choice == "secondhand" and pump_row_id:
            row = Seconhand.objects.get(id=pump_row_id)
            row.pump = repair.pump
            row.save()
            repair.pump_info = f"2.El Depo | {row.row_identifier}"

        #---------- Mütahit ------
        if engine_choice == "contractor" and engine_contractor_row_id and engine_row_id:
            row = Seconhand.objects.get(id=engine_row_id)
            row.engine = Engine.objects.get(id=engine_contractor_row_id)
            row.save()
            engine_locations_update(row.engine.id, "3")
            repair.engine_info = f"Mütahit Depo | {row.engine.serialnumber} | {row.row_identifier}"
            if repair.engine:
                engine_locations_update(repair.engine.id, "6")
        
        repair.status = 'passive'
        repair.save()
        order = repair.order
        if order.situation == "dismantling":
            order.operation_type = "5"
            messages.success(request, "Tamir bilgileri eklendi.\n Malzemler İlgili depoya aktarıldı.\n Montaj emri oluşturuldu.")
        elif order.situation == "well_cancellation":
            order.operation_type = "10"
            order.status ="passive"
            inv = order.inventory
            inv.engine = None
            inv.pump = None
            inv.status = "passive"
            inv.save()
            messages.success(request, "Tamir bilgileri eklendi.\n Malzemler İlgili depoya aktarıldı.\n Kuyu PASİF duruma aktarıldı.")
        else:
            order.operation_type = "10"
            order.status ="passive"
            messages.success(request, "Tamir bilgileri eklendi.\n Malzemler İlgili depoya aktarıldı.\n İş emri kapandı.")
        order.save()
        return redirect(next_url or "engine_repair")

    return render(request, "repair_edit.html", {
        "repair": repair,"engine":engine ,
        "row_identifier": row_identifier,
        'next': next_url
    })

@administrator
def repair_delete(request, id):
    return handle_deletion(
        request,
        Repair,
        id,
        'order_page',
        "*{0}* Tamir İş Emri kaydı başarıyla silindi.",
        "Tamir İş Emri kaydı bulunamadı.",
        "*{0}* Tamir İş Emri kaydı {1} tabloda kullanılıyor, silinemez."
    )
    
def contractor_warehouse(request):
    if request.method == "POST":
        form = EngineForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.location= "6" #Mütahit depo
            item.save()
            messages.success(request, f"Motor kaydı başarılı.")
            return redirect('contractor_warehouse')
        else:
            messages.warning(request, form.errors.as_ul())
    else:
        form = EngineForm()
    
    queryset = Engine.objects.filter(location="6")
    engine_filter = GeneralEngineFilter(request.GET, queryset=queryset)
    filtered_qs = engine_filter.qs
    page_obj = paginate_items(request, filtered_qs)
    
    contex = {
        'filter': engine_filter,
        'total': filtered_qs.count(),
        'items': page_obj, 
        'query_string': request.GET.urlencode(),
        'form' : form
    }
    return render(request, "contractor_warehouse.html",contex)

def unusable(request):
    queryset = Unusable.objects.all()
    engine_filter = UnusableFilter(request.GET, queryset=queryset)
    filtered_qs = engine_filter.qs
    page_obj = paginate_items(request, filtered_qs)
    
    contex = {
        'filter': engine_filter,
        'total': filtered_qs.count(),
        'items': page_obj, 
        'query_string': request.GET.urlencode(),
    }
    return render(request, "unusable_page.html",contex)

@admin
def export_unusable(request):
    # queryset ve filtreleme
    unusable_list = Unusable.objects.select_related('well_number', 'pump', 'engine').all()

    unusable_filter = UnusableFilter(request.GET, queryset=unusable_list)
    filtered_unusable_list = unusable_filter.qs

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Pert Depolar")

    # Excel başlıkları
    ws.append([
        "Kuyu Numarası",
        "Pompa",
        "Motor",
        "Oluşturulma Tarihi"
    ])

    # verileri iterator ile yaz
    for obj in filtered_unusable_list.iterator(chunk_size=1000):
        ws.append([
            obj.well_number.well_number if obj.well_number else "",
            str(obj.pump) if obj.pump else "",
            str(obj.engine) if obj.engine else "",
            obj.created_at.strftime("%Y-%m-%d %H:%M") if obj.created_at else "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="pert_depo.xlsx"'

    wb.save(response)
    return response

def new_warehouse_engine(request):
    if request.method == "POST":
        form = EngineForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f"Kayıt başarılı.")
            return redirect('new_warehouse_engine')
        else:
            messages.warning(request, form.errors.as_ul())
    else:
        form = EngineForm()
        
    queryset = Engine.objects.filter(location="5").order_by("-id")
    engine_filter = GeneralEngineFilter(request.GET, queryset=queryset)
    filtered_qs = engine_filter.qs
    page_obj = paginate_items(request, filtered_qs)
    
    context = {
        'filter': engine_filter,
        'total': filtered_qs.count(),
        'items': page_obj,  
        'query_string': request.GET.urlencode(),
        'form' : form,
        'null_list': Seconhand.objects.filter(engine__isnull=True,pump__isnull=True)

    }
    return render(request, "new_warehouse_engine.html",context)

@admin
def export_engine(request):
    # queryset ve filtreleme
    engine_list = Engine.objects.select_related('engine_power', 'engine_mark').all()

    engine_filter = GeneralEngineFilter(request.GET, queryset=engine_list)
    filtered_engine_list = engine_filter.qs

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Sıfır Motorlar")

    # Excel başlıkları
    ws.append([
        "Motor Tipi",
        "Motor Gücü (HP)",
        "Marka",
        "Seri Numarası",
        "Açıklama"
    ])

    # verileri iterator ile yaz
    for obj in filtered_engine_list.iterator(chunk_size=1000):
        ws.append([
            obj.get_engine_type_display(),
            obj.engine_power.engine_power if obj.engine_power else "",
            obj.engine_mark.engine_mark if obj.engine_mark else "",
            obj.serialnumber,
            obj.comment,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="sifir_motorlar.xlsx"'

    wb.save(response)
    return response

def new_warehouse_pump(request):
    if request.method == "POST":
        form = WarehousePumpForm(request.POST)
        if form.is_valid():
            item = form.save(False)
            if form.data["quantity_value"]:
                item.quantity = form.data["quantity"]
            item.save()
            messages.success(request, f"Kayıt başarılı.")
            return redirect('new_warehouse_pump')
        else:
            messages.warning(request, form.errors.as_ul())
    else:
        form = WarehousePumpForm()
    
    queryset = NewWarehousePump.objects.all()

    inventory_filter = NewPumpFilter(request.GET, queryset=queryset)
    
    filtered_qs = inventory_filter.qs

    page_obj = paginate_items(request, filtered_qs)

    context = {
        'filter': inventory_filter,
        'items': page_obj,
        'total': filtered_qs.count(),
        'query_string': request.GET.urlencode(),
        'form' : form,
        'null_list': Seconhand.objects.filter(engine__isnull=True,pump__isnull=True)
    }
    
    return render(request, "new_warehouse_pump.html",context)

@admin
def export_new_warehouse_pump(request):
    # queryset ve filtreleme
    pump_list = NewWarehousePump.objects.select_related('pump').all()

    pump_filter = NewPumpFilter(request.GET, queryset=pump_list)
    filtered_pump_list = pump_filter.qs

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Yeni Pompa")

    # Excel başlıkları
    ws.append([
        "Pompa",
        "Miktar",
    ])

    # verileri iterator ile yaz
    for obj in filtered_pump_list.iterator(chunk_size=1000):
        ws.append([
            str(obj.pump) if obj.pump else "",
            obj.quantity,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="yeni_pompa.xlsx"'

    wb.save(response)
    return response

@admin
def transfer_warehouse(request):
    if request.method == "POST":
        item = get_object_or_404(Seconhand,row_identifier=request.POST.get("row_identifier"))
        if request.POST.get("modal_value") == "pump":
            pump = get_object_or_404(NewWarehousePump,pump=request.POST.get("pump_value"))
            item.pump = pump.pump
            pump.quantity -=1
            pump.save()
            messages.success(request, f"*{request.POST.get('row_identifier')}* Sırasına {item.pump} Pompa Eklendi.")
            item.save()
            return redirect('new_warehouse_pump')
        elif request.POST.get("modal_value") == "engine":
            engine = get_object_or_404(Engine,pk=request.POST.get("engine_value"))
            item.engine = engine
            engine.location = "3"
            engine.save()
            messages.success(request, f"*{request.POST.get('row_identifier')}* Sırasına {item.engine} Motor Eklendi.")
            item.save()
        else:
            messages.warning(request, f"Değer Girilmedi..")
    return redirect('new_warehouse_engine')

@administrator
def new_warehouse_pump_delete(request, id):
    return handle_deletion(
        request,
        NewWarehousePump,
        id,
        'new_warehouse_pump',
        "*{0}* Pompa kaydı başarıyla silindi.",
        "Pompa kaydı bulunamadı.",
        "*{0}* Pompa kaydı {1} tabloda kullanılıyor, silinemez."
    )

#------------İş Emirleri----------
@admin
def export_order(request):
    # queryset ve filtreleme
    order_list = Order.objects.select_related(
        'inventory', 'mounted_engine', 'mounted_pump', 'disassembled_engine', 'disassembled_pump'
    ).all()

    # Status filtresi (aktif / pasif)
    status_filter = request.GET.get('status_filter')
    if status_filter == "active":
        order_list = order_list.filter(status="active")
    elif status_filter == "passive":
        order_list = order_list.filter(status="passive")

    order_filter = OrderFilter(request.GET, queryset=order_list)
    filtered_order_list = order_filter.qs

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="İş Emirleri")

    ws.append([
        "Kuyu Numarası",
        "İlçe",
        "Adres",
        "İş Emri Fişi",
        "Çıkış Fişi",
        "Demontaj Fişi",
        "Montaj Fişi",
        "Atölyeden Giden Fiş",
        "Montaj Edilen Motor",
        "Montaj Edilen Pompa",
        "Demontaj Edilen Motor",
        "Demontaj Edilen Pompa",
        "Çıkış Fişi Tarihi",
        "Atölyeden Giden Fiş Tarihi",
        "Durum",
        "Açıklama"
    ])

    for obj in filtered_order_list.iterator(chunk_size=1000):
        ws.append([
            obj.inventory.well_number if obj.inventory else "",
            obj.inventory.get_district_display() if obj.inventory else "",
            obj.inventory.address if obj.inventory else "",
            obj.work_order_plug,
            obj.outlet_plug,
            obj.disassembly_plug,
            obj.assembly_plug,
            obj.entrance_plug,
            str(obj.mounted_engine) if obj.mounted_engine else "",
            str(obj.mounted_pump) if obj.mounted_pump else "",
            str(obj.disassembled_engine) if obj.disassembled_engine else "",
            str(obj.disassembled_pump) if obj.disassembled_pump else "",
            obj.outlet_plug_date.strftime("%Y-%m-%d") if obj.outlet_plug_date else "",
            obj.entrance_plug_date.strftime("%Y-%m-%d") if obj.entrance_plug_date else "",
            obj.get_status_display(),
            obj.comment
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="is_emirleri.xlsx"'

    wb.save(response)
    return response

@administrator
def delete_order(request, id):
    return handle_deletion(
        request,
        Order,
        id,
        'order_page',
        "*E{0}* Nolu İş Emri kaydı başarıyla silindi.",
        "İş Emri kaydı bulunamadı.",
        "*{0}* İş Emri kaydı {1} tabloda kullanılıyor, silinemez."
    )

@admin
def form_control(request):#*
    if request.method != "POST":
        messages.warning(request, "Hatalı seçim yapıldı.")
        return redirect("order_page")

    order = get_object_or_404(Order, pk=request.POST["id"])

    if order.operation_type == "1": #Demontaj
        """1 -> 2"""
        form = DisassemblyForm(request.POST, instance=order)
        if form.is_valid():
            item = form.save()
            order.complete_disassembly(item.disassembly_plug)
            messages.success(request, "Demontaj bilgileri eklendi.\n Malzemler kurtarıcı depoya aktarıldı.")
        else:
            messages.warning(request, form.errors.as_ul())
            return redirect("order_page")

    elif order.operation_type == "2":
        """2 -> 3"""
        order.arrive_workshop()
        messages.success(request, "Kurtarıcı bilgileri eklendi.\n Malzemler Atölye depoya aktarıldı.")
    
    elif order.operation_type == "3":
        """3 -> 4"""
        form = OrderEditForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            order.start_repair()
            if order.disassembled_engine:
                engine_locations_update(order.disassembled_engine.id, "2")
            messages.success(request, "Atölye bilgileri eklendi.\n Malzemler Tamir depoya aktarıldı.")
        else:
            messages.warning(request, form.errors.as_ul())
            return redirect("order_page")
        
    elif order.operation_type == "6": #Montaj
        """6 -> 10"""
        form = MountingForm(request.POST, instance=order)
        if form.is_valid():
            item = form.save()
            order.complete_assembly(item.assembly_plug)
            create_workshop_exit_slip("diver",order)
            messages.success(request, "Montaj bilgileri eklendi.\n İş Emri Kapatıldı.")
        else:
            messages.warning(request, form.errors.as_ul())
            return redirect("order_page")

    elif order.operation_type == "8":
        form = LenghtForm(request.POST, instance=order)
        other = CategoryStockOut2Form(request.POST)
        if form.is_valid() and other.is_valid():
            item = form.save()
            stock =other.save(commit=False)
            stock.outlet_plug = item.outlet_plug
            stock.well_number = order.inventory.well_number
            stock.district = order.inventory.district
            stock.address = order.inventory.address
            stock.save()
            create_workshop_exit_slip("other",stock)
            order.complete_lenght(item.length)
            messages.success(request, "Boy ekleme bilgileri eklendi.\n İş Emri Kapandı.")
        else:
            messages.warning(request, form.errors.as_ul())
            messages.warning(request, other.errors.as_ul())
            return redirect("order_page")
        
    return redirect(request.META.get('HTTP_REFERER', 'order_page'))

def order_page(request):
    if request.method == "POST":
        well_number = request.POST.get("well_number").strip().upper()
        if well_number:
            if DebtSituation.objects.filter(inventory__well_number = well_number).exists():
                messages.warning(request, f"*{well_number}* Kuyu için borç sorgulaması yapılmakta. Eğer sorgulama bitmiş ise *BORÇ DURUMU* ekranına aktarıldınız burdan iş emrine aktarabilirsiniz.")
                url = reverse("debt_situation") + f"?well_number={well_number}"
                return redirect(url)
            
            try:
                well = Inventory.objects.get(well_number=well_number)
                active_order = Order.objects.filter(inventory=well, status="active")
                if len(active_order) > 2:
                    messages.info(request, "Bu kuyu numarasında 3 aktif iş emri mevcut.")
                else:
                    if well.status == "passive":
                        if len(active_order) > 0:
                            messages.info(request,"YENİ kuyu olduğundan dolayı ikinci iş emri açılamamakta.Mevcut iş emrini kapattıktan sonra işlemlere devam edebilirsiniz.")
                            return redirect("order_page")
                        messages.info(request,"Kuyu PASİF durumda motor ve pompa eklerseniz AKTİF duruma geçecektir.")
                    if len(active_order) > 0:
                        messages.info(request,f"Kuyuda AKTİF iş emri var. *{len(active_order)+1}*. İş emri açılacaktır. ")
                    return redirect("transactions", id=well.id)
            except Inventory.DoesNotExist:
                messages.warning(request, "Bu kuyu numarası bulunamadı.")
                return redirect("order_page")

    queryset = Order.objects.filter(status="active").select_related('inventory')
    order_filter = OrderFilter(request.GET, queryset=queryset)
    filtered_qs = order_filter.qs

    page_obj = paginate_items(request, filtered_qs)
    context = {
        'total': filtered_qs.count(),
        'items': page_obj,
        'query_string': request.GET.urlencode(),
        'filter': order_filter,
    }
    return render(request, "order_page.html", context)

def order_show(request, pk):
    order = get_object_or_404(Order, pk=pk)
    data = dict()
    repair = Repair.objects.filter(order=order).first()
    if request.POST:
        order.comment = request.POST.get('comment', None)
        order.save()
        messages.success(request,"İş emrine açıklama eklendi.")
        return redirect("order_page")
    context = {
        'order': order,
        'repair' : repair if repair else None,
        }
    data['html_form'] = render_to_string('modal/order_edit.html', context, request=request, )
    return JsonResponse(data)

@admin
def order_edit(request, pk):#*
    order = get_object_or_404(Order, pk=pk)
    data = {}
    context = {"order": order}

    if order.operation_type == "1":
        form = DisassemblyForm()
        context['form'] = form
        data['html_form'] = render_to_string('modal/disassembly.html', context, request=request)

    elif order.operation_type == "2":
        data['html_form'] = render_to_string('modal/workshop.html', context, request=request)

    elif order.operation_type == "3":
        form = OrderEditForm()
        context['form'] = form
        data['html_form'] = render_to_string('modal/disassembly.html', context, request=request)
    
    elif order.operation_type == "6":
        form = MountingForm()
        context['form'] = form
        data['html_form'] = render_to_string('modal/disassembly.html', context, request=request)

    elif order.operation_type == "8":
        form = LenghtForm()
        other = CategoryStockOut2Form()
        context['form'] = form
        context['other'] = other
        data['html_form'] = render_to_string('modal/lenght.html', context, request=request)

    return JsonResponse(data)

@transaction.atomic
def seconhand_order_go_back(request, id):
    order = get_object_or_404(Order, pk=id)

    if request.method == "POST":

        engine_row_id = request.POST.get("engine_secondhand_row")
        pump_row_id = request.POST.get("pump_secondhand_row")

        # ENGINE
        if engine_row_id and order.mounted_engine:
            try:
                item = Seconhand.objects.filter(id=engine_row_id).first()
                if item:
                    item.engine = order.mounted_engine
                    item.save()
                    engine_locations_update(item.engine.id, "3")
                else:
                    messages.warning(request, "Seçilen motor satırı bulunamadı.")
            except Exception as e:
                messages.error(request, f"Motor geri alınırken hata oluştu: {e}")

        # PUMP
        if pump_row_id and order.mounted_pump:
            try:
                item = Seconhand.objects.filter(id=pump_row_id).first()
                if item:
                    item.pump = order.mounted_pump
                    item.save()
                else:
                    messages.warning(request, "Seçilen pompa satırı bulunamadı.")
            except Exception as e:
                messages.error(request, f"Pompa geri alınırken hata oluştu: {e}")

        # Order temizleme
        order.mounted_engine = None
        order.mounted_pump = None
        order.outlet_plug = None
        order.operation_type = "5"
            
        order.save()
        messages.success(request,"İş emri geri alındı.")
        return redirect("order_page")

    return render(request, "seconhand_order_go_back.html", {
        "order": order,
        "row_identifier": Seconhand.objects.filter(
            engine__isnull=True,
            pump__isnull=True
        )
    })

@administrator
def order_go_back(request, pk):#*
    order = get_object_or_404(Order, pk=pk)
    if order.operation_type == "1":
        return redirect("seconhand_order_go_back",id=order.id)
    elif order.operation_type == "6" and ( order.situation == "dismantling" or order.situation == "new_well" ) :
        return redirect("seconhand_order_go_back",id=order.id)
        
    success, message = order.go_back()
    if success:
        messages.success(request, message)
    else:
        messages.warning(request, message)
    return redirect("order_page")

def handle_engine(order):
    if not order.mounted_engine:
        raise ValueError("Motor seçilmedi.")
    engine = Engine.objects.get(id=order.mounted_engine.id)
    if str(engine.location) == "3": 
        item = Seconhand.objects.get(engine=engine)
        order.engine_info = item.row_identifier
        item.engine = None
        item.save()
    engine.location = "1" 
    engine.save()

def handle_pump(order, pump_display):
    if not order.mounted_pump or not pump_display:
        raise ValueError("Pompa seçilmedi.")

    row_identifier = pump_display.split("-")
    if row_identifier[0].strip() == "Dalgıç":
        row_identifier = row_identifier[1].strip()
    else:
        row_identifier = row_identifier[0].strip()
    try:
        item = Seconhand.objects.get(
            pump_id=order.mounted_pump,
            row_identifier=row_identifier
        )
        order.pump_info = item.row_identifier
        item.pump=None
        item.save()

    except Seconhand.DoesNotExist:
        try:
            new_pump = NewWarehousePump.objects.get(pump=order.mounted_pump)
            if new_pump.quantity <= 0:
                raise ValueError("Seçilen pompa stoğu yok.")
            new_pump.quantity -= 1
            new_pump.save()
        except NewWarehousePump.DoesNotExist:
            raise ValueError("Seçilen pompa bulunamadı.")

@admin
def new_assembly_order(request, id):
    order = get_object_or_404(Order, id=id)
    if request.method == "POST":
        form = AssemblyForm(request.POST, instance=order)
        if form.is_valid():
            try:
                with transaction.atomic():
                    order = form.save(commit=False)
                    if order.operation_engine == "engine":
                        order.mounted_pump = None
                        handle_engine(order)
                        
                    elif order.operation_engine == "pump":
                        order.mounted_engine = None
                        handle_pump(order,request.POST.get("mounted_pump_display"))
                    else:
                        handle_engine(order)
                        handle_pump(order,request.POST.get("mounted_pump_display"))
                    if order.situation in ["new_well","dismantling"] :
                        order.operation_type = "6"
                        messages.success(request, " Çıkış Fişi bilgileri eklendi.\n Montaj bilgilerine aktarıldı.")
                    else:
                        order.operation_type = "1"
                        messages.success(request, "Malzemeler Kuyuya Gönderildi.")
                    order.save()
                    return redirect("order_page")

            except ValueError as ve:
                messages.error(request, f"Hata: {str(ve)}")
            except Exception as e:
                messages.error(request, f"İşlem sırasında beklenmedik hata oluştu: {str(e)}")
        else:
            messages.warning(request, form.errors.as_ul())
    else:
        form = AssemblyForm()

    return render(request, "new_order.html", {
        "order" :order,
        "form": form,
        "seconhand": Seconhand.objects.all(),
        "new_engine": Engine.objects.filter(location=5),
        "new_pump": NewWarehousePump.objects.filter(quantity__gt=0),
    })

def all_order_page(request):
    queryset = Order.objects.filter(status="passive").select_related('inventory').order_by("-id")

    order_filter = OrderFilter(request.GET, queryset=queryset)
    filtered_qs = order_filter.qs

    page_obj = paginate_items(request, filtered_qs)
    context = {
        'total': filtered_qs.count(),
        'items': page_obj,
        'query_string': request.GET.urlencode(),
        'filter': order_filter,
    }
    return render(request, "all_order_page.html", context)

@admin
def transactions(request,id, debt_id=None):
    inventory = get_object_or_404(Inventory, id=id)
    form = OperationForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            order = form.save(commit=False)
            order.inventory = inventory
            item = form.cleaned_data.get("situation")
            if item == "installation":
                order.operation_type = "5"
            elif item == "dismantling":
                order.operation_type = "1"
            elif item == "length_extension":
                order.operation_type = "8"
            elif item == "well_cancellation":
                order.operation_type = "1"
                order.operation_engine = "engine_pump"
            elif item == "new_well":
                order.operation_type = "5"
                order.operation_engine = "engine_pump"
                
            order.save()
            if debt_id:
                DebtSituation.objects.filter(id=debt_id, inventory=inventory).delete()
            messages.success(request, f"*{inventory.well_number}* Nolu kuyu için *E{order.work_order_plug}* iş emri oluşturuldu.")
            return redirect("order_page")
            
        else:
            messages.warning(request, form.errors.as_ul())

    return render(request, "transactions.html", {
        "form": form,
        "inventory": inventory,
        "end_order" : Order.objects.filter(inventory=inventory).order_by("id").first()
    })

def create_workshop_exit_slip(modalname, modal):
    if modalname == 'hydrophore':
        workshop_exit_slip = WorkshopExitSlip.objects.create(
            date=modal.dispatch_date,
            slip_no=modal.dispatch_slip_number,
            well_no="",
            district=modal.get_district_display(),
            address=modal.neighborhood,
            motor_type="HİDROFOR",
            hydrofor_no=modal.mounted_hydrophore.serial_number if modal.mounted_hydrophore else None,
            brand=modal.mounted_hydrophore.engine_brand if modal.mounted_hydrophore else None,
            power=modal.mounted_hydrophore.engine_power if modal.mounted_hydrophore else None,
            pump_type=modal.mounted_hydrophore.pump_type if modal.mounted_hydrophore else None,
            pump_brand="",
            submersible=None,
            motor=None,
            pump=None,
            hydrofor="1",
            overall_status="",
        )
    elif modalname == 'diver':
        workshop_exit_slip = WorkshopExitSlip.objects.create(
            date=modal.outlet_plug_date,
            slip_no=modal.outlet_plug,
            well_no=modal.inventory.well_number,
            district=modal.inventory.get_district_display(),
            address=modal.inventory.address,
            motor_type=modal.mounted_engine.get_engine_type_display() if modal.mounted_engine else None,
            hydrofor_no=modal.mounted_engine.serialnumber if modal.mounted_engine else None,
            brand=modal.mounted_engine.engine_mark if modal.mounted_engine else None,
            power=modal.mounted_engine.engine_power if modal.mounted_engine else None,
            pump_type=modal.mounted_pump.pump_type if modal.mounted_pump else None,
            pump_brand=modal.mounted_pump.pump_mark if modal.mounted_pump else None,
            submersible="1" if modal.mounted_engine and modal.mounted_pump else None,
            motor="2.EL-"+modal.engine_info if modal.engine_info else None,
            pump="2.EL-"+modal.pump_info if modal.pump_info else None,
            hydrofor=None,
            overall_status=modal.comment,
            modal_id = modalname + "/" + str(modal.id)
        )
    elif modalname == 'other':
        workshop_exit_slip = WorkshopExitSlip.objects.create(
            slip_no=modal.outlet_plug,
            date=modal.created_at,
            well_no=modal.well_number,
            district=modal.get_district_display(),
            address=modal.address,
            main_pipe = modal.stock,
            secondary_pipe = str(modal.quantity) + " Adet",
            maintenance_status = "Boy ekleme yapıldı.",
            modal_id = modalname + "/" + str(modal.id)
        )
    return

def workshop_exit_slip(request):
    order_list = WorkshopExitSlip.objects.all()
    
    order_filter = WorkshopExitSlipFilter(request.GET, queryset=order_list)
    filtered_order_list = order_filter.qs

    paginator = Paginator(filtered_order_list, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'total': filtered_order_list.count(),
        'items': page_obj,
        'query_string': request.GET.urlencode(),
        'filter': order_filter,
    }
    return render(request, "workshop_exit_slip.html", context)

@admin
def export_workshop_exit_slips(request):
    # filtreleri uygula
    order_list = WorkshopExitSlip.objects.all()
    order_filter = WorkshopExitSlipFilter(request.GET, queryset=order_list)
    filtered_order_list = order_filter.qs

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="atölye_cıkış_fişleri")

    # Excel Header
    ws.append([
        "Tarih",
        "Fiş No",
        "Kuyu No",
        "İlçe",
        "Adres",
        "Motor Tipi",
        "Hidrofor No",
        "Markası",
        "Gücü",
        "Pompa Tipi",
        "Pompa Markası",
        "Dalgıç",
        "Motor",
        "Pompa",
        "Hidrofor",
        "Ö.Boru",
        "K.Boru",
        "Bakım Durumu",
        "Genel Durum",
    ])

    # iterator ile bellek dostu export
    for obj in filtered_order_list.iterator(chunk_size=1000):
        ws.append([
            obj.date.strftime("%Y-%m-%d") if obj.date else "",
            obj.slip_no,
            obj.well_no,
            obj.district,
            obj.address,
            obj.motor_type,
            obj.hydrofor_no,
            obj.brand,
            obj.power,
            obj.pump_type,
            obj.pump_brand,
            obj.submersible,
            obj.motor,
            obj.pump,
            obj.hydrofor,
            obj.main_pipe,
            obj.secondary_pipe,
            obj.maintenance_status,
            obj.overall_status,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="atölye_cıkış_fişleri.xlsx"'

    wb.save(response)
    return response

@admin
def new_workshop_exit_slip(request):
    if request.method == 'POST':
        form = WorkshopExitSlipForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "İş Emri Eklendi.")
            return redirect('workshop_exit_slip')
        else:
            messages.warning(request, form.errors.as_ul())
    else:
            form = WorkshopExitSlipForm()
    return render(request, 'new_workshop_exit_slip.html', {'form': form})

@administrator
def workshop_exit_slip_edit(request,id):
    order = get_object_or_404(WorkshopExitSlip, pk=id)
    if request.method == 'POST':
        form = WorkshopExitSlipForm(request.POST,instance=order)
        if form.is_valid():
            form.save()
            messages.success(request, "İş Emri Güncellendi.")
            return redirect('workshop_exit_slip')
        else:
            messages.warning(request, form.errors.as_ul())
    else:
            form = WorkshopExitSlipForm(instance=order)
    return render(request, 'new_workshop_exit_slip.html', {'form': form})

@administrator
def workshop_exit_slip_delete(request, id):
    return handle_deletion(
        request=request,
        model=WorkshopExitSlip,
        object_id=id,
        redirect_url="workshop_exit_slip",
        success_message="{} Kuyusu için iş emri başarıyla silindi.",
        error_message="İş emri bulunamadı.",
        protected_error_message="*{0}* İş Emri kaydı {1} tabloda kullanılıyor, silinemez."
    )

from other_materials.models import CategoryStockOut
def work_order_reporting(request):
    order = None
    hydrophore = None
    repair = None
    workshop = None
    repair_return = None
    other_stock = None

    if request.method == 'POST':
        item = request.POST.get("number")

        order = Order.objects.filter(outlet_plug=item).first()
        hydrophore = OutboundWorkOrder.objects.filter(dispatch_slip_number=item).first()
        other_stock = CategoryStockOut.objects.filter(outlet_plug=item)
        
        if order:
            repair = Repair.objects.filter(order=order).first()

        elif hydrophore:
            workshop = WorkshopExit.objects.filter(outbound_work_order=hydrophore).first()
            if workshop:
                repair_return = RepairReturn.objects.filter(workshop_exit=workshop).first()

        elif other_stock:
           other_stock = other_stock
        else:
            messages.warning(request, "Çıkış fiş numarası bulunamadı.")
    context = {
        'other_stock' : other_stock, 
        'order': order,
        'hydrophore': hydrophore,
        'repair': repair,
        'workshop': workshop,
        'repair_return': repair_return,
    }

    return render(request, 'work_order_reporting.html', context)

#Borç Durum sorgulama
def debt_situation(request):
    if request.method == "POST":
        well_number = request.POST.get("well_number").strip().upper()
        if well_number:
            if DebtSituation.objects.filter(inventory__well_number = well_number).exists():
                messages.warning(request, f"*{well_number}* Kuyu kaydı mevcut.")
                return redirect(debt_situation)
                
            well = Inventory.objects.filter(well_number=well_number)
            if well.exists():
                return redirect("new_debt_situation", id=well[0].id)
            else: 
                messages.warning(request, "Bu kuyu numarası bulunamadı.")
                return redirect("debt_situation")

    queryset = DebtSituation.objects.all().select_related('inventory')
    order_filter = DebtSituationFilter(request.GET, queryset=queryset)
    filtered_qs = order_filter.qs

    page_obj = paginate_items(request, filtered_qs)
    context = {
        'total': filtered_qs.count(),
        'items': page_obj,
        'query_string': request.GET.urlencode(),
        'filter': order_filter,
    }
    return render(request, "debt_situation.html", context)

@admin
def new_debt_situation(request,id):
    inventory = get_object_or_404(Inventory, id=id)
    if request.method == "POST":
        form = DebtSituationForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.inventory = inventory
            item.save()
            print(item.inventory)
            messages.success(request, f"*{inventory.well_number}* Kuyu için borç sorgulama kaydı açıldı.")
            return redirect("debt_situation")
        else:
            messages.warning(request, form.errors.as_ul())
    else:
        form = DebtSituationForm()

    return render(request, "new_debt_situation.html", {
        "inventory" :inventory,
        "form": form,
    })

@administrator
def delete_debt_situation(request, id):
    return handle_deletion(
        request,
        DebtSituation,
        id,
        'debt_situation',
        "*{0}* Borç durumu başarıyla silindi.",
        "Borç durumu bulunamadı.",
        "*{0}* Borç durumu kaydı {1} tabloda kullanılıyor, silinemez."
    )

def transfer_debt_situation(request, id):
    debt = get_object_or_404(DebtSituation, id=id)
    active_order = Order.objects.filter(inventory=debt.inventory, status="active")

    if active_order.count() > 2:
        messages.info(request, "Bu kuyu numarasında 3 aktif iş emri mevcut. İş emri aktarılamadı.")
        return redirect("debt_situation")

    else:
        if debt.inventory.status == "passive":
            if active_order.count() > 0:
                messages.info(
                    request,
                    "YENİ kuyu olduğundan dolayı ikinci iş emri açılamamakta. Mevcut iş emrini kapattıktan sonra işlemlere devam edebilirsiniz."
                )
                return redirect("debt_situation")

            messages.info(
                request,
                "Kuyu PASİF durumda motor ve pompa eklerseniz AKTİF duruma geçecektir."
            )

        if active_order.count() > 0:
            messages.info(
                request,
                f"Kuyuda AKTİF iş emri var. *{active_order.count()+1}.* İş emri açılacaktır."
            )

    return redirect("transactions", id=debt.inventory.id, debt_id=debt.id)
